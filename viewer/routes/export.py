"""
viewer/routes/export.py — Blueprint Flask pour les exports et le chatbot.

Routes :
  GET        /api/export/atom
  GET/POST   /api/export/newsletter
  POST       /api/export/webhook-test
  GET        /api/export/csv
  GET        /api/export/xlsx
  POST       /api/chat/stream       (streaming SSE)
  POST       /api/chat/save
"""
import datetime
import hashlib
import json
import os
import re

from flask import Blueprint, jsonify, request, Response, stream_with_context, abort
from pathlib import Path

from viewer.helpers import PROJECT_ROOT
from viewer.state import _annotations_lock
from utils.article_index import get_article_index
from utils.scoring import get_scoring_engine

export_bp = Blueprint("export", __name__)

# Paramètres du chatbot
_CHAT_MAX_CONTEXT_FILES   = 10     # Nombre maximum de fichiers de contexte par requête
_CHAT_MAX_CONTEXT_CHARS   = 100000 # Taille maximale (caractères) par fichier de contexte (brut)
_CHAT_MAX_ARTICLES_JSON   = 100    # Nombre max d'articles inclus depuis un fichier JSON d'articles
_CHAT_MAX_ENTITIES_TYPE   = 5      # Nombre max d'entités par type dans le contexte article compact


def _load_annotations_for_chat() -> dict:
    """Charge le fichier annotations.json pour le chatbot."""
    annotations_file = PROJECT_ROOT / "data" / "annotations.json"
    if not annotations_file.exists():
        return {}
    try:
        return json.loads(annotations_file.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _build_notes_context(period: str = "week") -> str:
    """Génère un bloc de contexte Markdown à partir des annotations personnelles."""
    annotations = _load_annotations_for_chat()
    if not annotations:
        return ""

    # Déterminer la date de début selon la période
    now = datetime.datetime.now(datetime.timezone.utc)
    if period == "week":
        cutoff = now - datetime.timedelta(days=7)
        period_label = "7 derniers jours"
    elif period == "month":
        cutoff = now - datetime.timedelta(days=30)
        period_label = "30 derniers jours"
    else:
        cutoff = None
        period_label = "toutes les notes"

    # Filtrer les annotations ayant une note ou des tags, et selon la période
    selected = {}
    for url, ann in annotations.items():
        has_content = ann.get("notes", "").strip() or [t for t in (ann.get("tags") or []) if t]
        if not has_content:
            continue
        if cutoff is not None:
            updated_raw = ann.get("updated_at", "")
            if updated_raw:
                try:
                    updated = datetime.datetime.fromisoformat(updated_raw.replace("Z", "+00:00"))
                    if updated.tzinfo is None:
                        updated = updated.replace(tzinfo=datetime.timezone.utc)
                    if updated < cutoff:
                        continue
                except Exception:
                    pass
        selected[url] = ann

    if not selected:
        return f"*Aucune note personnelle pour la période : {period_label}.*"

    # Construire un index article {url: article_dict} pour enrichir avec le titre/source
    article_index: dict = {}

    def _index_articles(directory: Path) -> None:
        if not directory.exists():
            return
        for f in directory.rglob("*.json"):
            parts = f.relative_to(PROJECT_ROOT).parts
            if "cache" in parts:
                continue
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    for a in data:
                        if isinstance(a, dict) and a.get("URL"):
                            article_index[a["URL"]] = a
            except Exception:
                pass

    _index_articles(PROJECT_ROOT / "data" / "articles-from-rss")
    _index_articles(PROJECT_ROOT / "data" / "articles")

    # Trier par date de mise à jour décroissante
    sorted_entries = sorted(
        selected.items(),
        key=lambda kv: kv[1].get("updated_at", ""),
        reverse=True,
    )

    lines = [
        f"## Notes personnelles de lecture ({period_label})",
        f"*{len(sorted_entries)} note(s) trouvée(s)*",
        "",
    ]

    for url, ann in sorted_entries:
        notes = ann.get("notes", "").strip()
        tags = [t for t in (ann.get("tags") or []) if t]
        is_important = bool(ann.get("is_important", False))
        updated_at = (ann.get("updated_at") or "")[:10]
        article = article_index.get(url, {})
        source = article.get("Sources", "")
        pub_date = (article.get("Date de publication") or "")[:10]

        # Extraire le titre de l'article
        titre = (article.get("Titre") or "").strip()
        if not titre:
            for line in (article.get("Résumé") or "").split("\n"):
                line = line.strip().lstrip("*_#").strip()
                if len(line) > 10:
                    titre = line[:120]
                    break
        if not titre:
            titre = url.rstrip("/").split("/")[-1][:80] or "Sans titre"

        star = "⭐ " if is_important else ""
        meta_parts = []
        if pub_date:
            meta_parts.append(pub_date)
        if source:
            meta_parts.append(source)
        meta = " · ".join(meta_parts)

        lines.append(f"### {star}{titre}")
        if meta:
            lines.append(f"*{meta}*")
        lines.append(f"URL : {url}")
        if updated_at:
            lines.append(f"Note ajoutée le : {updated_at}")
        if tags:
            lines.append(f"Tags : {', '.join(tags)}")
        if notes:
            lines.append(f"\n> {notes}")
        lines.append("")

    return "\n".join(lines)


def _format_articles_for_context(articles: list, max_articles: int = 100) -> str:
    """Formate une liste d'articles de manière compacte pour le contexte du chatbot.

    Au lieu d'inclure le JSON brut (verbeux), cette fonction génère un texte
    structuré Markdown qui inclut tous les articles sans troncature arbitraire.
    Chaque article occupe ~500-800 caractères, soit 5 à 10 fois moins que le
    JSON brut équivalent.

    Args:
        articles:     Liste de dicts article (format WUDD.ai).
        max_articles: Nombre maximum d'articles à inclure.

    Returns:
        Texte Markdown prêt à être inséré dans le contexte système.
    """
    total = len(articles)
    shown = min(total, max_articles)
    lines: list[str] = [
        f"*{total} article(s) au total — {shown} inclus dans le contexte.*",
        "",
    ]
    for i, art in enumerate(articles[:max_articles], 1):
        date    = art.get("Date de publication", "?")
        src     = art.get("Sources", "?")
        url     = art.get("URL", "")
        resume  = (art.get("Résumé") or "").strip()
        sentiment = art.get("sentiment", "")
        entities  = art.get("entities", {})

        lines.append(f"**Article {i}** — {date} | {src}")
        if url:
            lines.append(f"URL : {url}")
        if sentiment:
            lines.append(f"Sentiment : {sentiment}")
        # Entités clés (PERSON, ORG, GPE uniquement pour rester compact)
        if isinstance(entities, dict):
            ent_parts = []
            for etype in ("PERSON", "ORG", "GPE"):
                vals = entities.get(etype, [])
                if isinstance(vals, list) and vals:
                    ent_parts.append(f"{etype}: {', '.join(vals[:_CHAT_MAX_ENTITIES_TYPE])}")
            if ent_parts:
                lines.append("Entités : " + " | ".join(ent_parts))
        if resume:
            lines.append(resume)
        lines.append("")

    if total > max_articles:
        lines.append(
            f"*…{total - max_articles} article(s) supplémentaire(s) non inclus "
            f"(limite fixée à {max_articles}).*"
        )

    return "\n".join(lines)


@export_bp.route("/api/export/atom")
def api_export_atom():
    """Génère et retourne un flux Atom pour un flux ou tous les articles.

    Paramètres :
      flux        : nom du flux (ex: "Intelligence-artificielle")
      keyword     : mot-clé (ex: "OpenAI")
      max_entries : nombre max d'entrées (défaut: 50)
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.exporters.atom_feed import generate_atom_feed, generate_atom_from_flux

        flux = request.args.get("flux", "").strip()
        keyword = request.args.get("keyword", "").strip()
        max_entries = min(int(request.args.get("max_entries", 50)), 200)

        # URL canonique dynamique
        actual_self_url = request.url

        if flux:
            xml = generate_atom_from_flux(PROJECT_ROOT, flux, max_entries, self_url=actual_self_url)
        elif keyword:
            kw_file = PROJECT_ROOT / "data" / "articles-from-rss" / f"{keyword}.json"
            if not kw_file.exists():
                return jsonify({"error": "Fichier keyword introuvable"}), 404
            articles = json.loads(kw_file.read_text(encoding="utf-8"))
            articles.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
            from utils.exporters.atom_feed import _FEED_ID_BASE
            xml = generate_atom_feed(
                articles, feed_title=f"WUDD.ai · {keyword}",
                feed_id=f"{_FEED_ID_BASE}keyword-{keyword.lower()}",
                self_url=actual_self_url,
                max_entries=max_entries,
            )
        else:
            # Tout agréger via article_index (2 semaines)
            all_articles = []
            seen_urls_atom: set = set()
            try:
                aidx = get_article_index(PROJECT_ROOT)
                recent_entries = aidx.get_recent(hours=336)  # 14 jours
                files_atom: dict[str, Path] = {}
                for entry in recent_entries:
                    rel = entry.get("file", "")
                    if rel and rel not in files_atom:
                        files_atom[rel] = PROJECT_ROOT / rel
                for rel, jf in files_atom.items():
                    try:
                        data = json.loads(jf.read_text(encoding="utf-8"))
                        if isinstance(data, list):
                            for a in data:
                                url = a.get("URL", "")
                                if url and url in seen_urls_atom:
                                    continue
                                if url:
                                    seen_urls_atom.add(url)
                                all_articles.append(a)
                    except Exception:
                        continue
            except Exception:
                # Fallback rglob
                for d in [PROJECT_ROOT / "data" / "articles", PROJECT_ROOT / "data" / "articles-from-rss"]:
                    if not d.exists():
                        continue
                    for jf in sorted(d.rglob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)[:10]:
                        if "cache" in str(jf):
                            continue
                        try:
                            data = json.loads(jf.read_text(encoding="utf-8"))
                            if isinstance(data, list):
                                all_articles.extend(data)
                        except Exception:
                            continue
            all_articles.sort(key=lambda a: a.get("Date de publication", ""), reverse=True)
            xml = generate_atom_feed(all_articles, feed_title="WUDD.ai · Veille complète",
                                     self_url=actual_self_url,
                                     max_entries=max_entries)

        return Response(xml, mimetype="application/atom+xml; charset=utf-8")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@export_bp.route("/api/export/newsletter", methods=["GET", "POST"])
def api_export_newsletter():
    """Génère une newsletter HTML depuis les articles récents.

    GET  → retourne le HTML brut
    POST → { send: true } pour envoyer par SMTP (si configuré)

    Paramètres GET :
      hours : fenêtre temporelle (défaut: 48)
      title : titre de la newsletter
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.exporters.newsletter import generate_newsletter_html, send_newsletter

        hours = int(request.args.get("hours", 48))
        title = request.args.get("title", "").strip() or \
            f"Veille WUDD.ai — {datetime.datetime.now().strftime('%d %B %Y')}"

        engine = get_scoring_engine(PROJECT_ROOT)
        articles = engine.get_top_articles(top_n=20, hours=hours)
        html = generate_newsletter_html(articles, title=title)

        # Sauvegarde locale
        nl_dir = PROJECT_ROOT / "rapports" / "html"
        nl_dir.mkdir(parents=True, exist_ok=True)
        slug = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        nl_file = nl_dir / f"newsletter_{slug}.html"
        nl_file.write_text(html, encoding="utf-8")

        if request.method == "POST":
            data = request.get_json(force=True) or {}
            if data.get("send"):
                success = send_newsletter(html, subject=title)
                return jsonify({"ok": success, "path": str(nl_file.relative_to(PROJECT_ROOT))})

        return Response(html, mimetype="text/html; charset=utf-8")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@export_bp.route("/api/export/webhook-test", methods=["POST"])
def api_webhook_test():
    """Teste l'envoi webhook avec les alertes actuelles.

    Body JSON : { platform: "discord"|"slack"|"ntfy"|"all" }
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from utils.exporters.webhook import send_discord, send_slack, send_ntfy, notify_alerts

        platform = (request.get_json(force=True) or {}).get("platform", "all")
        alerts_file = PROJECT_ROOT / "data" / "alertes.json"
        alerts = []
        if alerts_file.exists():
            try:
                alerts = json.loads(alerts_file.read_text(encoding="utf-8"))
            except Exception:
                pass

        if not alerts:
            return jsonify({"ok": False, "message": "Aucune alerte disponible — lancez d'abord trend_detector.py"})

        if platform == "discord":
            ok = send_discord(alerts)
            return jsonify({"discord": ok})
        elif platform == "slack":
            ok = send_slack(alerts)
            return jsonify({"slack": ok})
        elif platform == "ntfy":
            ok = send_ntfy(alerts)
            return jsonify({"ntfy": ok})
        else:
            results = notify_alerts(alerts)
            return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@export_bp.route("/api/export/csv")
def api_export_csv():
    """Exporte un fichier JSON d'articles en CSV.

    Paramètre :
      path : chemin relatif du fichier JSON (ex: data/articles-from-rss/OpenAI.json)
    """
    import csv
    import io
    from viewer.helpers import safe_path

    path = request.args.get("path", "").strip()
    if not path:
        abort(400, "Paramètre path requis")
    f = safe_path(path)

    try:
        articles = json.loads(f.read_text(encoding="utf-8", errors="replace"))
        if not isinstance(articles, list):
            abort(400, "Le fichier ne contient pas une liste d'articles")
    except json.JSONDecodeError as e:
        abort(400, f"JSON invalide : {e}")

    # Colonnes exportées
    FIELDS = ["Date de publication", "Sources", "URL", "Résumé",
              "sentiment", "score_sentiment", "ton_editorial", "score_ton", "score_pertinence"]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=FIELDS, extrasaction="ignore",
                             lineterminator="\n")
    writer.writeheader()
    for art in articles:
        row = {k: art.get(k, "") for k in FIELDS}
        # Aplatir les entités en chaîne
        entities = art.get("entities", {})
        if isinstance(entities, dict):
            row["entities"] = "; ".join(
                f"{etype}:{','.join(str(v) for v in vals)}"
                for etype, vals in entities.items() if isinstance(vals, list)
            )
        writer.writerow(row)

    csv_content = output.getvalue()
    stem = Path(path).stem
    return Response(
        csv_content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'},
    )


@export_bp.route("/api/export/xlsx")
def api_export_xlsx():
    """Exporte un fichier JSON d'articles en XLSX (Excel).

    Paramètre :
      path : chemin relatif du fichier JSON
    """
    from viewer.helpers import safe_path

    path = request.args.get("path", "").strip()
    if not path:
        abort(400, "Paramètre path requis")
    f = safe_path(path)

    try:
        articles = json.loads(f.read_text(encoding="utf-8", errors="replace"))
        if not isinstance(articles, list):
            abort(400, "Le fichier ne contient pas une liste d'articles")
    except json.JSONDecodeError as e:
        abort(400, f"JSON invalide : {e}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"

        FIELDS = ["Date de publication", "Sources", "URL", "Résumé",
                  "sentiment", "score_sentiment", "ton_editorial", "score_ton",
                  "score_pertinence", "Entités"]

        # En-tête
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="FFFFFF")
        for col, field in enumerate(FIELDS, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.fill = header_fill
            cell.font = header_font

        # Données
        for row_idx, art in enumerate(articles, 2):
            entities = art.get("entities", {})
            entity_str = ""
            if isinstance(entities, dict):
                entity_str = "; ".join(
                    f"{et}:{','.join(str(v) for v in vals)}"
                    for et, vals in entities.items() if isinstance(vals, list)
                )
            values = [
                art.get("Date de publication", ""),
                art.get("Sources", ""),
                art.get("URL", ""),
                art.get("Résumé", ""),
                art.get("sentiment", ""),
                art.get("score_sentiment", ""),
                art.get("ton_editorial", ""),
                art.get("score_ton", ""),
                art.get("score_pertinence", ""),
                entity_str,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if col == 4:  # Résumé
                    cell.alignment = Alignment(wrap_text=True)

        # Largeurs de colonnes
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 80
        ws.row_dimensions[1].height = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        stem = Path(path).stem
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'},
        )
    except ImportError:
        # Fallback : retourne un CSV si openpyxl n'est pas installé
        return api_export_csv()
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@export_bp.route("/api/chat/stream", methods=["POST"])
def api_chat_stream():
    """Chatbot IA en streaming SSE.

    Body JSON :
      messages      (list)   — historique de conversation [{ role, content }, ...]
      context_files (list)   — chemins relatifs des fichiers à inclure comme contexte (optionnel)
      notes_period  (string) — période des notes personnelles à inclure : "week", "month" ou "all" (optionnel)

    Retourne un flux SSE au format OpenAI : data: {"choices":[{"delta":{"content":"..."},...}]}
    """
    import requests as req

    body = request.get_json(force=True, silent=True) or {}
    messages        = body.get("messages", [])
    context_files   = body.get("context_files", [])
    notes_period    = body.get("notes_period", None)
    # Contexte entité pré-formaté (texte brut) fourni par Terminal IA depuis EntityArticlePanel
    entity_context  = body.get("entity_context", "").strip()
    # Permet au frontend de choisir le provider pour cette requête.
    provider_override = body.get("provider", "").strip().lower()

    if not messages:
        return jsonify({"error": "messages est requis"}), 400

    # Valider et charger les fichiers de contexte
    context_blocks = []
    for rel in context_files[:_CHAT_MAX_CONTEXT_FILES]:
        rel = str(rel).strip()
        if not rel:
            continue
        # Restriction aux répertoires autorisés
        if not (rel.startswith("data/") or rel.startswith("rapports/") or rel.startswith("samples/")):
            continue
        target = (PROJECT_ROOT / rel).resolve()
        if not str(target).startswith(str(PROJECT_ROOT) + "/"):
            continue
        if not target.exists() or not target.is_file():
            continue
        try:
            raw = target.read_text(encoding="utf-8")
            ext = target.suffix.lower()
            lang = "json" if ext == ".json" else "markdown"

            # Extraction intelligente pour les fichiers JSON d'articles :
            # au lieu de tronquer le JSON brut (ce qui limite le contexte à ~5
            # articles), on parse le tableau et on formate chaque article de
            # manière compacte, ce qui inclut jusqu'à _CHAT_MAX_ARTICLES_JSON
            # articles dans un espace ~5× plus petit.
            if ext == ".json":
                try:
                    data = json.loads(raw)
                    if (
                        isinstance(data, list)
                        and data
                        and isinstance(data[0], dict)
                        and "Résumé" in data[0]
                    ):
                        formatted = _format_articles_for_context(data, _CHAT_MAX_ARTICLES_JSON)
                        context_blocks.append(f"### Fichier : {rel}\n{formatted}")
                        continue
                except (json.JSONDecodeError, IndexError, KeyError):
                    pass  # Fallback : traitement générique ci-dessous

            # Fallback générique : tronquer les fichiers trop volumineux
            if len(raw) > _CHAT_MAX_CONTEXT_CHARS:
                raw = raw[:_CHAT_MAX_CONTEXT_CHARS] + "\n…[tronqué]"
            context_blocks.append(f"### Fichier : {rel}\n```{lang}\n{raw}\n```")
        except OSError:
            continue

    # Charger les notes personnelles si demandé
    notes_block = None
    if notes_period and notes_period in ("week", "month", "all"):
        with _annotations_lock:
            notes_text = _build_notes_context(notes_period)
        if notes_text:
            notes_block = notes_text

    # Construire le message système
    from datetime import datetime as _dt
    _today = _dt.now().strftime("%A %d %B %Y")
    system_parts = [
        "Tu es un assistant IA intégré à WUDD.ai, une plateforme de veille de presse en français.",
        f"La date d'aujourd'hui est le {_today}. Tiens-en compte pour contextualiser toutes tes réponses sur l'actualité.",
        "Tu aides l'utilisateur à analyser des articles de presse, des rapports et des données JSON.",
        "Tu peux produire des tableaux Markdown, des résumés, des analyses comparatives.",
        "Réponds toujours en français, de manière concise et structurée.",
        "Utilise du Markdown pour les tableaux, listes et mise en forme.",
        "Ne génère pas de balises <think>.",
        "IMPORTANT — Tu es un assistant en LECTURE SEULE. Tu ne peux PAS supprimer, effacer, modifier ou détruire des fichiers, des données ou des rapports.",
        "Si l'utilisateur te demande de supprimer ou d'effacer des fichiers, des données ou des rapports"
        " (quelle que soit la formulation : commandes shell, appels API, code, instructions, etc.),"
        " refuse poliment et rappelle-lui que cette opération est impossible depuis ce chatbot.",
    ]
    if entity_context:
        system_parts.append("\n\n## Contexte entité (données WUDD.ai) :\n")
        system_parts.append(entity_context)
    if notes_block:
        system_parts.append("\n\n## Notes personnelles de l'utilisateur :\n")
        system_parts.append(notes_block)
    if context_blocks:
        system_parts.append("\n\n## Fichiers de contexte fournis par l'utilisateur :\n")
        system_parts.extend(context_blocks)

    system_prompt = "\n".join(system_parts)

    # Nettoyer les messages
    clean_messages = [
        {"role": m["role"], "content": m["content"]}
        for m in messages
        if m.get("role") in ("user", "assistant") and (m.get("content") or "").strip()
    ]

    # Construire les messages complets (système + historique)
    full_messages = [{"role": "system", "content": system_prompt}] + clean_messages

    provider = provider_override if provider_override in ("euria", "claude") \
               else os.environ.get("AI_PROVIDER", "euria").strip().lower()

    if provider == "claude":
        api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY manquante dans .env (AI_PROVIDER=claude)"}), 503
        # Routage Haiku/Sonnet selon la taille totale du contexte (système + historique).
        # On inclut le system_prompt car il peut contenir des fichiers d'articles volumineux.
        total_context_chars = len(system_prompt) + sum(len(m.get("content", "")) for m in clean_messages)
        if total_context_chars < 3000:
            model = os.environ.get("CLAUDE_MODEL_BATCH", "claude-haiku-4-5-20251001")
        else:
            model = os.environ.get("CLAUDE_MODEL_SYNTHESIS", "claude-sonnet-4-6")
        # Claude ne supporte pas role=system dans messages[], on extrait le system
        claude_messages = [m for m in full_messages if m["role"] != "system"]
        from utils.api_client import ClaudeClient as _ClaudeClient
        _claude = _ClaudeClient(api_key=api_key)

        def generate_chat():
            yield from _claude.stream(
                prompt="",
                model=model,
                system=system_prompt,
                max_tokens=4096,
                timeout=180,
                messages=claude_messages,
            )

    else:
        api_url = os.environ.get("URL", "").strip()
        bearer  = os.environ.get("bearer", "").strip()
        if not api_url or not bearer:
            return jsonify({"error": "URL ou bearer manquant dans .env (AI_PROVIDER=euria)"}), 503
        payload = {
            "messages": full_messages,
            "model": "qwen3",
            "stream": True,
            "enable_web_search": True,
        }
        api_headers = {
            "Authorization": f"Bearer {bearer}",
            "Content-Type": "application/json",
        }

        def generate_chat():
            try:
                r = req.post(api_url, json=payload, headers=api_headers, stream=True, timeout=180)
                r.raise_for_status()
                for line in r.iter_lines():
                    if line:
                        decoded = line.decode("utf-8")
                        if not decoded.startswith("data:"):
                            decoded = "data: " + decoded
                        yield decoded + "\n\n"
            except Exception as exc:
                yield f'data: {json.dumps({"error": str(exc)})}\n\n'

    return Response(
        stream_with_context(generate_chat()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@export_bp.route("/api/export/report", methods=["POST"])
def api_export_report():
    """Sauvegarde un rapport Markdown sur le filesystem local ou dans le répertoire Obsidian.

    Body JSON :
      markdown  (str)  — contenu Markdown du rapport
      filename  (str)  — nom de fichier suggéré (sans extension)
      target    (str)  — 'local' (rapports/markdown/_WUDD.AI_/) ou 'obsidian' (OBSIDIAN_DIR)
      resume    (str)  — résumé de l'article (utilisé pour la déduplication MD5, optionnel)

    Retourne : { ok: bool, path: str, deduplicated?: bool }
    """
    body     = request.get_json(force=True, silent=True) or {}
    markdown = (body.get("markdown") or "").strip()
    filename = (body.get("filename") or "rapport").strip()
    target   = (body.get("target") or "local").strip().lower()
    resume   = (body.get("resume") or "").strip()   # pour déduplication

    if not markdown:
        return jsonify({"error": "markdown est requis"}), 400

    if target == "obsidian":
        obsidian_dir = os.environ.get("OBSIDIAN_DIR", "").strip()
        if not obsidian_dir:
            return jsonify({"error": "OBSIDIAN_DIR non configuré dans .env"}), 400
        save_dir  = Path(obsidian_dir)
        # Obsidian : le frontend formate déjà le nom (YYYY-MM-DD_source_slug.md)
        safe_name = re.sub(r"[^\w\-]", "_", filename)[:120] + ".md"
    else:
        save_dir = PROJECT_ROOT / "rapports" / "markdown" / "_WUDD.AI_"
        # Export local : ajouter un timestamp pour éviter les collisions
        fname_clean = re.sub(r"[^\w\-]", "_", filename)[:80]
        ts          = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name   = f"{fname_clean}_{ts}.md"

    try:
        save_dir.mkdir(parents=True, exist_ok=True)

        # ── Déduplication par MD5 du résumé (Obsidian uniquement) ─────────────
        if target == "obsidian" and resume:
            resume_md5 = hashlib.md5(resume.encode("utf-8")).hexdigest()
            index_path = save_dir / ".wudd_export_index.json"
            index: dict = {}
            if index_path.exists():
                try:
                    index = json.loads(index_path.read_text(encoding="utf-8"))
                except Exception:
                    index = {}
            if resume_md5 in index:
                existing = Path(index[resume_md5])
                if existing.exists():
                    saved_at_dedup = datetime.datetime.fromtimestamp(
                        existing.stat().st_mtime
                    ).isoformat(timespec="seconds")
                    return jsonify({
                        "ok": True,
                        "path": str(existing),
                        "filename": existing.name,
                        "saved_at": saved_at_dedup,
                        "deduplicated": True,
                    })
        else:
            resume_md5 = None
            index_path = None
            index      = {}

        out_path = save_dir / safe_name
        out_path.write_text(markdown, encoding="utf-8")

        # Mettre à jour l'index de déduplication
        if target == "obsidian" and resume_md5 and index_path is not None:
            index[resume_md5] = str(out_path)
            try:
                index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")
            except OSError:
                pass  # l'index est optionnel, ne pas bloquer l'export

        saved_at = datetime.datetime.now().isoformat(timespec="seconds")
        return jsonify({
            "ok": True,
            "path": str(out_path),
            "filename": safe_name,
            "saved_at": saved_at,
        })
    except OSError as e:
        return jsonify({"error": f"Erreur écriture : {e}"}), 500


@export_bp.route("/api/article/set-report-meta", methods=["POST"])
def api_article_set_report_meta():
    """Enregistre les métadonnées d'un rapport exporté directement dans l'article JSON.

    Body JSON :
      file_path    (str)  — chemin absolu du fichier JSON d'articles
      article_url  (str)  — URL de l'article à mettre à jour
      rapport      (obj)  — { fichier, chemin, cible, date_creation }

    Retourne : { ok: bool }
    """
    body        = request.get_json(force=True, silent=True) or {}
    file_path   = (body.get("file_path") or "").strip()
    article_url = (body.get("article_url") or "").strip()
    rapport     = body.get("rapport") or {}

    if not file_path or not article_url or not rapport:
        return jsonify({"error": "file_path, article_url et rapport sont requis"}), 400

    # Résolution : chemin relatif à PROJECT_ROOT ou chemin absolu
    fp = Path(file_path)
    if not fp.is_absolute():
        fp = PROJECT_ROOT / fp
    fp = fp.resolve()

    if not fp.exists() or not fp.is_file():
        return jsonify({"error": "Fichier introuvable"}), 404

    # Sécurité : le fichier doit être sous PROJECT_ROOT
    try:
        fp.resolve().relative_to(PROJECT_ROOT.resolve())
    except ValueError:
        return jsonify({"error": "Accès refusé"}), 403

    try:
        articles = json.loads(fp.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": f"Lecture impossible : {e}"}), 500

    if not isinstance(articles, list):
        return jsonify({"error": "Format JSON invalide (liste attendue)"}), 400

    updated = False
    for art in articles:
        if art.get("URL") == article_url:
            rapports = art.get("rapports") or []
            # Éviter les doublons : même fichier
            if not any(r.get("fichier") == rapport.get("fichier") for r in rapports):
                rapports.append(rapport)
            art["rapports"] = rapports
            updated = True
            break

    if not updated:
        return jsonify({"error": "Article introuvable dans le fichier"}), 404

    try:
        fp.write_text(json.dumps(articles, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError as e:
        return jsonify({"error": f"Écriture impossible : {e}"}), 500

    # Propager le rapport aux fichiers sources (articles-from-rss/*.json et
    # articles/**/*.json) pour que les métadonnées survivent à la reconstruction
    # de la fenêtre glissante 48-heures.json.
    _propagate_rapport_to_source_files(article_url, rapport, fp, PROJECT_ROOT)

    return jsonify({"ok": True})


def _propagate_rapport_to_source_files(
    article_url: str,
    rapport: dict,
    updated_file: Path,
    project_root: Path,
) -> None:
    """Propage les métadonnées de rapport aux fichiers sources contenant l'article.

    Recherche l'article par URL dans les fichiers JSON sources (articles-from-rss/
    et articles/<flux>/) et y ajoute le rapport si absent.  N'échoue jamais de
    façon bloquante : les erreurs sont ignorées silencieusement.

    Args:
        article_url  : URL de l'article mis à jour.
        rapport      : Métadonnées du rapport { fichier, chemin, cible, date_creation }.
        updated_file : Fichier déjà mis à jour (éviter de le réécrire une 2e fois).
        project_root : Racine du projet WUDD.ai.
    """
    fichier_rapport = rapport.get("fichier", "")
    scan_dirs = [
        project_root / "data" / "articles-from-rss",
        project_root / "data" / "articles",
    ]
    for scan_dir in scan_dirs:
        if not scan_dir.exists():
            continue
        # articles-from-rss : fichiers .json au niveau racine du répertoire
        # articles : fichiers .json dans les sous-répertoires de flux
        pattern = "*.json" if scan_dir.name == "articles-from-rss" else "**/*.json"
        for json_file in scan_dir.glob(pattern):
            if json_file.resolve() == updated_file.resolve():
                continue
            if "cache" in json_file.parts:
                continue
            try:
                other_articles = json.loads(json_file.read_text(encoding="utf-8"))
                if not isinstance(other_articles, list):
                    continue
                other_changed = False
                for other_art in other_articles:
                    if other_art.get("URL") == article_url:
                        other_rapports = other_art.get("rapports") or []
                        if not any(
                            r.get("fichier") == fichier_rapport for r in other_rapports
                        ):
                            other_rapports.append(rapport)
                            other_art["rapports"] = other_rapports
                            other_changed = True
                        break  # Premier article correspondant à l'URL traité
                if other_changed:
                    json_file.write_text(
                        json.dumps(other_articles, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
            except Exception:
                continue  # propagation optionnelle, ne jamais bloquer


@export_bp.route("/api/entity/set-report-meta", methods=["POST"])
def api_entity_set_report_meta():
    """Enregistre les métadonnées d'un rapport d'entité dans l'index partagé.

    Body JSON :
      entity_type  (str)  — type NER (ORG, PERSON, GPE, …)
      entity_value (str)  — valeur de l'entité
      rapport      (obj)  — { fichier, chemin, cible, date_creation }

    Retourne : { ok: bool }
    """
    body         = request.get_json(force=True, silent=True) or {}
    entity_type  = (body.get("entity_type") or "").strip()
    entity_value = (body.get("entity_value") or "").strip()
    rapport      = body.get("rapport") or {}

    if not entity_type or not entity_value or not rapport:
        return jsonify({"error": "entity_type, entity_value et rapport sont requis"}), 400

    index_path = PROJECT_ROOT / "data" / "entity_reports_index.json"
    index: dict = {}
    if index_path.exists():
        try:
            index = json.loads(index_path.read_text(encoding="utf-8"))
        except Exception:
            index = {}

    key = f"{entity_type}:{entity_value}"
    rapports = index.get(key) or []
    if not any(r.get("fichier") == rapport.get("fichier") for r in rapports):
        rapports.append(rapport)
    index[key] = rapports

    try:
        index_path.parent.mkdir(parents=True, exist_ok=True)
        index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError as e:
        return jsonify({"error": f"Écriture impossible : {e}"}), 500

    return jsonify({"ok": True})


@export_bp.route("/api/entity/get-report-meta", methods=["GET"])
def api_entity_get_report_meta():
    """Retourne les métadonnées de rapports pour une entité.

    Query params :
      entity_type  (str)
      entity_value (str)

    Retourne : { rapports: [...] }
    """
    entity_type  = request.args.get("entity_type", "").strip()
    entity_value = request.args.get("entity_value", "").strip()

    if not entity_type or not entity_value:
        return jsonify({"rapports": []})

    index_path = PROJECT_ROOT / "data" / "entity_reports_index.json"
    if not index_path.exists():
        return jsonify({"rapports": []})

    try:
        index = json.loads(index_path.read_text(encoding="utf-8"))
    except Exception:
        return jsonify({"rapports": []})

    key = f"{entity_type}:{entity_value}"
    return jsonify({"rapports": index.get(key, [])})


@export_bp.route("/api/config/obsidian", methods=["GET"])
def api_config_obsidian():
    """Retourne la configuration Obsidian pour construire les liens obsidian://.

    vault_name est retourné uniquement si OBSIDIAN_VAULT_NAME est explicitement
    défini dans .env. Ne pas le dériver automatiquement depuis le chemin :
    le nom Obsidian peut différer du nom de répertoire (accents, espaces, etc.)

    Retourne : { vault_name: str | null, obsidian_dir: str | null }
    """
    obsidian_dir = os.environ.get("OBSIDIAN_DIR", "").strip() or None
    vault_name   = os.environ.get("OBSIDIAN_VAULT_NAME", "").strip() or None

    return jsonify({"vault_name": vault_name, "obsidian_dir": obsidian_dir})


@export_bp.route("/api/chat/save", methods=["POST"])
def api_chat_save():
    """Sauvegarde une conversation ou une réponse IA en Markdown dans rapports/.

    Body JSON :
      content  (str)  — contenu Markdown à sauvegarder
      filename (str)  — nom de fichier suggéré (sans extension, optionnel)
      subdir   (str)  — sous-répertoire dans rapports/ (défaut : "_WUDD.AI_")

    Retourne : { ok: bool, path: str }
    """
    body = request.get_json(force=True, silent=True) or {}
    content  = (body.get("content") or "").strip()
    filename = (body.get("filename") or "").strip()
    subdir   = (body.get("subdir") or "_WUDD.AI_").strip()

    if not content:
        return jsonify({"error": "content est requis"}), 400

    # Sanitiser le sous-répertoire
    subdir = re.sub(r"[^\w\-/]", "_", subdir).strip("/")
    if not subdir:
        subdir = "_WUDD.AI_"

    # Sanitiser le nom de fichier
    if filename:
        filename = re.sub(r"[^\w\-]", "_", filename)[:80]
    else:
        filename = "chat"

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"{filename}_{ts}.md"

    save_dir = (PROJECT_ROOT / "rapports" / subdir).resolve()
    # Vérifier que le répertoire cible reste dans rapports/
    rapports_root = (PROJECT_ROOT / "rapports").resolve()
    if not str(save_dir).startswith(str(rapports_root)):
        return jsonify({"error": "Répertoire non autorisé"}), 403

    try:
        save_dir.mkdir(parents=True, exist_ok=True)
        out_path = save_dir / safe_name
        out_path.write_text(content, encoding="utf-8")
        rel = str(out_path.relative_to(PROJECT_ROOT)).replace("\\", "/")
        return jsonify({"ok": True, "path": rel})
    except OSError as e:
        return jsonify({"error": f"Erreur écriture : {e}"}), 500
